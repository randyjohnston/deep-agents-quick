"""In-process XLSX read/write for the agent.

An .xlsx file is a ZIP archive of Office Open XML parts (workbook.xml, the
per-sheet XML, styles.xml, sharedStrings.xml, and their .rels). Building or
parsing one needs no sandbox and no subprocess — openpyxl does it in memory, so
this works under `langgraph dev` and in a deployed container alike.

Files live on the real filesystem rather than the agent's virtual StateBackend:
state files are stored as text, so binary written there is base64-encoded and
unopenable. Shared Office path policy confines reads and writes and rejects
oversized OOXML ZIP packages before openpyxl parses them.
"""

from __future__ import annotations

import re
from datetime import date, datetime, time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from app.tools.office.paths import resolve_read_path, resolve_write_path

CellValue = bool | int | float | str | None

MAX_ROWS = 1_048_576
MAX_COLS = 16_384

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_SHEET_NAME_LIMIT = 31

_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_ISO_DATETIME = re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}(:\d{2})?$")

# A leading one of these makes Excel treat text as a formula.
_FORMULA_LEAD = ("=", "+", "-", "@")

_COL_WIDTH_MIN = 8
_COL_WIDTH_MAX = 60

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")


class Sheet(BaseModel):
    """One worksheet: an optional header row plus the data rows below it."""

    name: str = Field(description="Worksheet tab name")
    columns: list[str] | None = Field(
        default=None, description="Header row; omit if the data has no headers"
    )
    rows: list[list[CellValue]] = Field(
        default_factory=list, description="Data rows, each a list of cell values"
    )


def _unique_sheet_name(raw: str, taken: set[str]) -> str:
    """Sanitize to Excel's rules, then de-duplicate with a numeric suffix."""
    name = _INVALID_SHEET_CHARS.sub("-", raw).strip("'").strip() or "Sheet"
    name = name[:_SHEET_NAME_LIMIT]

    candidate, n = name, 2
    while candidate.casefold() in taken:
        suffix = f"_{n}"
        candidate = f"{name[: _SHEET_NAME_LIMIT - len(suffix)]}{suffix}"
        n += 1
    taken.add(candidate.casefold())
    return candidate


def _coerce_date(value: str) -> datetime | date | None:
    """Promote strict ISO-8601 strings to real Excel dates.

    Deliberately narrow: JSON has no date type, so date-like strings would
    otherwise land as text and break sorting and date math. Anything not
    exactly ISO (a bare year, an ID like "2024-01") stays a string.
    """
    matcher = _ISO_DATE if _ISO_DATE.match(value) else _ISO_DATETIME
    if not matcher.match(value):
        return None
    try:
        return date.fromisoformat(value) if matcher is _ISO_DATE else datetime.fromisoformat(value)
    except ValueError:
        return None


def _set_cell(ws, row: int, col: int, value: CellValue) -> int:
    """Write one cell with its natural Excel type; return its rendered width."""
    if value is None:
        return 0

    cell = ws.cell(row=row, column=col)
    if isinstance(value, bool):
        cell.value = value
        return len(str(value))
    if isinstance(value, (int, float)):
        cell.value = value
        return len(str(value))

    text = str(value)
    if (coerced := _coerce_date(text)) is not None:
        cell.value = coerced
        if isinstance(coerced, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
            return 19
        cell.number_format = "yyyy-mm-dd"
        return 10

    cell.value = text
    if text.startswith(_FORMULA_LEAD):
        # openpyxl would store this as a live formula; force inline string so
        # untrusted data cannot become executable (XLSX/CSV injection).
        cell.data_type = "s"
    return len(text)


def write_xlsx(
    filename: str,
    sheets: list[Sheet],
    freeze_header: bool = True,
    autofilter: bool = True,
) -> str:
    """Write data to a real .xlsx workbook and return its path.

    Use this instead of a CSV when the user asks for Excel/XLSX output, when the
    result needs multiple tabs, or when types (dates, numbers) must survive the
    round trip. Each sheet gets a bold frozen header, an autofilter, and column
    widths sized to the content.

    Args:
        filename: Output basename, e.g. "customers.xlsx". Directories are
            rejected; the file is written under OFFICE_OUTPUT_DIR.
        sheets: One entry per worksheet.
        freeze_header: Keep the header row visible while scrolling.
        autofilter: Add filter dropdowns across the header row.

    Returns:
        A summary naming the absolute path and the shape of each sheet written.
    """
    if not sheets:
        raise ValueError("Provide at least one sheet.")

    path = resolve_write_path(filename, ".xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; we add our own

    taken: set[str] = set()
    summary: list[str] = []

    for sheet in sheets:
        header = sheet.columns or []
        width_count = max([len(header), *(len(r) for r in sheet.rows)] or [0])
        total_rows = len(sheet.rows) + (1 if header else 0)
        if total_rows > MAX_ROWS:
            raise ValueError(
                f"Sheet {sheet.name!r} has {total_rows} rows; Excel allows {MAX_ROWS}."
            )
        if width_count > MAX_COLS:
            raise ValueError(
                f"Sheet {sheet.name!r} has {width_count} columns; Excel allows {MAX_COLS}."
            )

        ws = wb.create_sheet(_unique_sheet_name(sheet.name, taken))
        widths = [0] * width_count

        if header:
            for col, title in enumerate(header, start=1):
                cell = ws.cell(row=1, column=col, value=str(title))
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                widths[col - 1] = len(str(title))

        row_offset = 2 if header else 1
        for r, row in enumerate(sheet.rows, start=row_offset):
            for c, value in enumerate(row, start=1):
                rendered = _set_cell(ws, r, c, value)
                widths[c - 1] = max(widths[c - 1], rendered)

        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = min(
                max(width + 2, _COL_WIDTH_MIN), _COL_WIDTH_MAX
            )

        if header:
            if freeze_header:
                ws.freeze_panes = "A2"
            if autofilter and sheet.rows and width_count:
                last = get_column_letter(width_count)
                ws.auto_filter.ref = f"A1:{last}{len(sheet.rows) + 1}"

        summary.append(f"{ws.title} ({len(sheet.rows)} rows x {width_count} cols)")

    wb.save(path)
    return f"Wrote {path} — " + "; ".join(summary)


def read_xlsx(
    path: str,
    sheet: str | None = None,
    max_rows: int = 200,
) -> str:
    """Read an existing .xlsx workbook and return its contents as text.

    Use this to inspect a spreadsheet the user supplied before transforming it.
    Formula cells come back as their last cached value, not the formula.

    Args:
        path: Workbook filename or path, resolved under OFFICE_INPUT_DIR or
            OFFICE_OUTPUT_DIR.
        sheet: Read only this worksheet; omit to read them all.
        max_rows: Row cap per sheet, so a large workbook cannot flood context.

    Returns:
        A pipe-delimited rendering of each sheet, noting any truncation.
    """
    resolved = resolve_read_path(path, (".xlsx",))

    # data_only surfaces cached formula results instead of "=SUM(...)".
    wb = load_workbook(resolved, read_only=True, data_only=True)
    try:
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise ValueError(
                    f"No sheet {sheet!r} in {resolved.name}. Available: {', '.join(wb.sheetnames)}"
                )
            names = [sheet]
        else:
            names = list(wb.sheetnames)

        blocks = [f"# {resolved}", f"Sheets: {', '.join(wb.sheetnames)}", ""]
        for name in names:
            ws = wb[name]
            blocks.append(f"## {name}")

            emitted = 0
            truncated = False
            for row in ws.iter_rows(values_only=True):
                if emitted >= max_rows:
                    truncated = True
                    break
                if all(v is None for v in row):
                    continue
                blocks.append(" | ".join(_render(v) for v in row))
                emitted += 1

            if emitted == 0:
                blocks.append("(empty)")
            if truncated:
                blocks.append(f"... truncated at {max_rows} rows; raise max_rows for more")
            blocks.append("")
    finally:
        wb.close()

    return "\n".join(blocks).rstrip()


def _render(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return str(value)
