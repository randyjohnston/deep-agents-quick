"""Round-trip tests: everything written is read back with openpyxl."""

from __future__ import annotations

from datetime import date, datetime

import pytest
from openpyxl import load_workbook

from xlsx import Sheet, read_xlsx, write_xlsx


@pytest.fixture(autouse=True)
def _isolate_dirs(tmp_path, monkeypatch):
    monkeypatch.setenv("XLSX_OUTPUT_DIR", str(tmp_path / "out"))
    monkeypatch.setenv("XLSX_INPUT_DIR", str(tmp_path / "in"))
    (tmp_path / "in").mkdir()
    return tmp_path


def _written(tmp_path, name="book.xlsx"):
    return tmp_path / "out" / name


def test_writes_openable_workbook_with_multiple_sheets(_isolate_dirs):
    result = write_xlsx(
        "report",
        [
            Sheet(name="Summary", columns=["Metric", "Value"], rows=[["Accounts", 412]]),
            Sheet(name="Detail", columns=["Account"], rows=[["Acme"], ["Globex"]]),
        ],
    )

    path = _written(_isolate_dirs, "report.xlsx")
    assert path.is_file(), result
    wb = load_workbook(path)
    assert wb.sheetnames == ["Summary", "Detail"]
    assert wb["Detail"].max_row == 3  # header + 2 rows


def test_extension_is_added_when_missing(_isolate_dirs):
    write_xlsx("noext", [Sheet(name="S", rows=[["x"]])])
    assert _written(_isolate_dirs, "noext.xlsx").is_file()


def test_types_survive_the_round_trip(_isolate_dirs):
    write_xlsx(
        "typed",
        [
            Sheet(
                name="T",
                columns=["num", "date", "stamp", "flag", "text", "blank"],
                rows=[[84000, "2026-03-14", "2026-03-14 09:30:00", True, "hello", None]],
            )
        ],
    )

    ws = load_workbook(_written(_isolate_dirs, "typed.xlsx"))["T"]
    assert ws["A2"].value == 84000
    assert ws["B2"].value == datetime(2026, 3, 14)
    assert ws["C2"].value == datetime(2026, 3, 14, 9, 30)
    assert ws["D2"].value is True
    assert ws["E2"].value == "hello"
    assert ws["F2"].value is None


@pytest.mark.parametrize("value", ["2026", "2026-01", "not-a-date", "14/03/2026"])
def test_non_iso_strings_stay_text(_isolate_dirs, value):
    write_xlsx("dates", [Sheet(name="D", columns=["v"], rows=[[value]])])
    ws = load_workbook(_written(_isolate_dirs, "dates.xlsx"))["D"]
    assert ws["A2"].value == value
    assert not isinstance(ws["A2"].value, (date, datetime))


@pytest.mark.parametrize("payload", ["=SUM(1,2)", "+1+1", "-1+1", "@SUM(A1)"])
def test_formula_like_text_is_not_stored_as_a_formula(_isolate_dirs, payload):
    write_xlsx("inject", [Sheet(name="I", columns=["v"], rows=[[payload]])])
    ws = load_workbook(_written(_isolate_dirs, "inject.xlsx"))["I"]
    assert ws["A2"].value == payload
    assert ws["A2"].data_type == "s", "must be inline text, not an executable formula"


def test_header_is_frozen_and_filtered(_isolate_dirs):
    write_xlsx("fmt", [Sheet(name="F", columns=["a", "b"], rows=[[1, 2]])])
    ws = load_workbook(_written(_isolate_dirs, "fmt.xlsx"))["F"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:B2"
    assert ws["A1"].font.bold


def test_sheet_names_are_sanitized_truncated_and_deduped(_isolate_dirs):
    write_xlsx(
        "names",
        [
            Sheet(name="Bad/Name:With*Chars", rows=[["x"]]),
            Sheet(name="dup", rows=[["x"]]),
            Sheet(name="dup", rows=[["x"]]),
            Sheet(name="x" * 40, rows=[["x"]]),
        ],
    )
    names = load_workbook(_written(_isolate_dirs, "names.xlsx")).sheetnames
    assert names[0] == "Bad-Name-With-Chars"
    assert names[1] == "dup" and names[2] == "dup_2"
    assert len(names[3]) == 31


def test_ragged_rows_do_not_lose_columns(_isolate_dirs):
    write_xlsx("ragged", [Sheet(name="R", columns=["a"], rows=[[1, 2, 3]])])
    ws = load_workbook(_written(_isolate_dirs, "ragged.xlsx"))["R"]
    assert [c.value for c in ws[2]] == [1, 2, 3]


def test_write_refuses_to_escape_the_output_root(_isolate_dirs):
    write_xlsx("../escaped.xlsx", [Sheet(name="S", rows=[["x"]])])
    # The basename is kept, so the file lands inside the root, not above it.
    assert _written(_isolate_dirs, "escaped.xlsx").is_file()
    assert not (_isolate_dirs / "escaped.xlsx").exists()


def test_empty_sheet_list_is_rejected(_isolate_dirs):
    with pytest.raises(ValueError, match="at least one sheet"):
        write_xlsx("empty", [])


def test_read_round_trips_what_write_produced(_isolate_dirs):
    write_xlsx(
        "rt",
        [Sheet(name="Data", columns=["Account", "ARR"], rows=[["Acme", 84000]])],
    )
    out = read_xlsx("rt.xlsx")
    assert "Data" in out
    assert "Acme | 84000" in out


def test_read_can_select_one_sheet(_isolate_dirs):
    write_xlsx(
        "multi",
        [
            Sheet(name="One", columns=["a"], rows=[["first"]]),
            Sheet(name="Two", columns=["b"], rows=[["second"]]),
        ],
    )
    out = read_xlsx("multi.xlsx", sheet="Two")
    assert "second" in out
    assert "first" not in out


def test_read_rejects_unknown_sheet(_isolate_dirs):
    write_xlsx("one", [Sheet(name="Only", columns=["a"], rows=[["v"]])])
    with pytest.raises(ValueError, match="No sheet 'Nope'"):
        read_xlsx("one.xlsx", sheet="Nope")


def test_read_truncates_at_max_rows(_isolate_dirs):
    write_xlsx(
        "big",
        [Sheet(name="B", columns=["n"], rows=[[i] for i in range(50)])],
    )
    out = read_xlsx("big.xlsx", max_rows=10)
    assert "truncated at 10 rows" in out


def test_read_reports_missing_file(_isolate_dirs):
    with pytest.raises(FileNotFoundError, match="No workbook found"):
        read_xlsx("absent.xlsx")


def test_read_finds_files_in_the_input_dir(_isolate_dirs):
    write_xlsx("seed", [Sheet(name="S", columns=["a"], rows=[["from-input"]])])
    src = _written(_isolate_dirs, "seed.xlsx")
    src.rename(_isolate_dirs / "in" / "seed.xlsx")

    assert "from-input" in read_xlsx("seed.xlsx")


def test_read_refuses_paths_outside_the_permitted_roots(_isolate_dirs):
    outside = _isolate_dirs / "secret.xlsx"
    write_xlsx("secret", [Sheet(name="S", rows=[["x"]])])
    _written(_isolate_dirs, "secret.xlsx").rename(outside)

    with pytest.raises(FileNotFoundError):
        read_xlsx(str(outside))
